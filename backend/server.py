import asyncio
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
import redis.asyncio as aioredis
from fastapi.middleware.cors import CORSMiddleware
from database import *

from datetime import datetime
from openpyxl import Workbook
from os import mkdir, path

from logger import Logger

log = Logger("server")
log.info("Started.")

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

db = Database(user="server", logger=log)

@app.get("/")
async def index():
    log.info("Request to /")
    return "WeatherStation API"

@app.get("/sse")
async def sse_endpoint():
    log.info("Request to /sse")

    async def event_generator():
        r = aioredis.Redis(host='localhost', port=6379, db=0)
        pubsub = r.pubsub()
        log.info("Subscribed to sse.")
        await pubsub.subscribe('live')
        
        try:
            async for message in pubsub.listen():
                if message['type'] == 'message':
                    data = message['data'].decode('utf-8')
                    # Pushes event to client instantly when Redis receives it
                    yield f"data:{data}\n\n"
        finally:
            log.info("Unsubscribe from sse.")
            await pubsub.unsubscribe('weather_updates')

    return StreamingResponse(event_generator(), media_type="text/event-stream")

@app.get("/export_excel")
def export_excel(dateFrom: str, dateTo: str, sensors: str):
    log.info("Request to /export_excel. Received args: dateFrom:"+dateFrom+"& dateTo:"+dateTo+"& sensors:"+sensors)
    # TODO: sql injection risk.
    data = db.get_weather_data(dateFrom, dateTo, sensors.split(","))
    if data == None:
        log.error("No data found in database for export.")
        raise HTTPException(status_code=422, detail="No data within specified timeframe.") 

    # Format the data into a spreadsheet
    wb = Workbook()
    for date in data:
        if date == "columns":
            continue

        ws = wb.create_sheet(date)

        c = 1
        for title in data["columns"]:
            if title == "precipitation":
                continue
            ws.cell(column=c, row=1, value=title)
            c += 1
        # precipitation is always moved to end
        ws.cell(column=c, row=1, value="precipitation")

        r = 2
        for row in data[date]:
            c = 1
            for col in row:
                ws.cell(column=c, row=r, value=col)
                c += 1
            r += 1

    # write spreadsheet to tmp (deleted on reboot)
    if not path.exists("/tmp/weatherstation"):
        mkdir("/tmp/weatherstation/")

    fpath = "/tmp/weatherstation/weatherstation_export_"+datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+".xlsx"
    log.debug("Export file written to "+fpath)
    wb.save(fpath)

    log.info("Export success")
    return str(fpath.split("/")[-1])

@app.get("/get_file/")
def get_file(file: str):
    file = file.replace("\"", "")

    log.info("Returned file "+"/tmp/weatherstation/"+file)
    # return files from temp
    return FileResponse("/tmp/weatherstation/"+file, filename=file)